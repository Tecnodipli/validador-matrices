import os
import re
import io
import uuid
import unicodedata
import zipfile
from collections import Counter
from datetime import datetime, timedelta
from io import BytesIO

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from docx import Document

# ==========================
# Configuración
# ==========================
app = FastAPI(title="Validador de Matrices")

ALLOWED_ORIGINS = [
    "https://www.dipli.ai",
    "https://dipli.ai",
    "https://isagarcivill09.wixsite.com/turop",
    "https://isagarcivill09.wixsite.com/turop/tienda",
    "https://isagarcivill09-wixsite-com.filesusr.com"
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==========================
# Descargas temporales
# ==========================
DOWNLOADS: dict[str, dict] = {}
EXP_MINUTES = 5  # tiempo de expiración del link

def cleanup_downloads():
    now = datetime.utcnow()
    expired = [t for t, v in DOWNLOADS.items() if v["exp"] <= now]
    for t in expired:
        DOWNLOADS.pop(t, None)

def register_download(data: bytes, filename: str, media_type: str) -> str:
    cleanup_downloads()
    token = str(uuid.uuid4())
    expires_at = datetime.utcnow() + timedelta(minutes=EXP_MINUTES)
    DOWNLOADS[token] = {
        "data": data,
        "filename": filename,
        "media_type": media_type,
        "exp": expires_at
    }
    return token

# ==========================
# Validaciones
# ==========================
CARACTERES_PROHIBIDOS = set("!@#$%&/()=\u00a1\u00a8*[];:_°|\u00ac")
ENCABEZADOS_ESPERADOS = ["Capitulo", "Subcapitulo", "Preguntas"]

def char_human(ch: str) -> str:
    code = f"U+{ord(ch):04X}"
    name = unicodedata.name(ch, "UNKNOWN")
    visible = ch if not ch.isspace() else repr(ch)
    return f"{visible} ({code} {name})"

def validar_encabezados(sheet):
    errores = []
    for col, esperado in zip(['A', 'B', 'C'], ENCABEZADOS_ESPERADOS):
        celda = f"{col}1"
        valor = str(sheet[celda].value).strip() if sheet[celda].value else ""
        if valor != esperado:
            errores.append(f"❌ Celda {celda} debería contener '{esperado}', pero tiene '{valor}'")
    return errores

def buscar_preguntas_duplicadas(sheet):
    from collections import defaultdict
    preguntas = defaultdict(list)
    for row in range(2, sheet.max_row + 1):
        valor = sheet[f"C{row}"].value
        if valor:
            valor = str(valor).strip()
            preguntas[valor].append(row)
    return [f"❌ Pregunta duplicada en filas {v}: '{k}'" for k, v in preguntas.items() if len(v) > 1]

def buscar_caracteres_prohibidos(sheet):
    errores = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for c in cell.value:
                    if c in CARACTERES_PROHIBIDOS:
                        errores.append(f"❌ Celda {cell.coordinate} contiene caracter prohibido '{c}' en: '{cell.value}'")
                        break
    return errores

# ==========================
# Endpoints
# ==========================
@app.post("/procesar/")
async def procesar(file: UploadFile = File(...)):
    import openpyxl

    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx")

    try:
        wb = openpyxl.load_workbook(file.file)
        hoja = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo abrir el archivo: {e}")

    errores = []
    errores.extend(validar_encabezados(hoja))
    errores.extend(buscar_preguntas_duplicadas(hoja))
    errores.extend(buscar_caracteres_prohibidos(hoja))

    # Crear reporte TXT en memoria
    txt_bytes = BytesIO()
    if not errores:
        txt_bytes.write("✅ VALIDACIÓN EXITOSA: No se encontraron errores.\n".encode("utf-8"))
    else:
        txt_bytes.write("❌ VALIDACIÓN FALLIDA: Se encontraron errores:\n\n".encode("utf-8"))
        for err in errores:
            txt_bytes.write(f"{err}\n".encode("utf-8"))
    txt_bytes.seek(0)

    # Registrar para descarga
    final_name = f"reporte_errores_{os.path.splitext(file.filename)[0]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    token = register_download(txt_bytes.getvalue(), final_name, "text/plain; charset=utf-8")

    return JSONResponse({"token": token, "filename": final_name})

@app.get("/download/{token}")
def download_token(token: str):
    cleanup_downloads()
    item = DOWNLOADS.get(token)
    if not item:
        raise HTTPException(status_code=404, detail="Link expirado o inválido")

    if item["exp"] <= datetime.utcnow():
        DOWNLOADS.pop(token, None)
        raise HTTPException(status_code=410, detail="Link expirado")

    headers = {
        "Content-Disposition": f'attachment; filename="{item["filename"]}"',
        "Cache-Control": "no-store"
    }
    return StreamingResponse(io.BytesIO(item["data"]), media_type=item["media_type"], headers=headers)

@app.get("/")
async def root():
    return {"message": "API de validación de matrices funcionando 🚀"}

@app.get("/health")
async def health_check():
    return {"status": "healthy"}
